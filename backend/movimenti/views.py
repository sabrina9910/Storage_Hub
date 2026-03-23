from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q, Sum

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from core.permissions import IsMagazziniere
from .models import StockMovement
from .serializers import StockMovementSerializer
from .filters import StockMovementFilter
from prodotti.models import Product, ProductLot
from users.models import User
from auditlog.utils import create_audit_log

class StockMovementViewSet(viewsets.ModelViewSet):
    serializer_class = StockMovementSerializer
    permission_classes = [IsMagazziniere]
    filterset_class = StockMovementFilter

    def get_queryset(self):
        user = self.request.user
        if not user or not user.is_authenticated:
            return StockMovement.objects.none()
        
        qs = StockMovement.objects.all().select_related(
            'product', 'product__category', 
            'lot', 'lot__product', 'lot__product__category', 
            'user'
        ).prefetch_related('product__suppliers', 'lot__product__suppliers')
        
        # Period filtering
        period = self.request.query_params.get('period')
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        movement_type = self.request.query_params.get('type')
        user_id = self.request.query_params.get('user')
        
        if period:
            now = timezone.now()
            if period == 'today':
                start = now.replace(hour=0, minute=0, second=0, microsecond=0)
                qs = qs.filter(timestamp__gte=start)
            elif period == 'week':
                start = now - timedelta(days=now.weekday())
                start = start.replace(hour=0, minute=0, second=0, microsecond=0)
                qs = qs.filter(timestamp__gte=start)
            elif period == 'month':
                start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                qs = qs.filter(timestamp__gte=start)
            elif period == 'year':
                start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
                qs = qs.filter(timestamp__gte=start)
        
        if date_from and date_to:
            qs = qs.filter(timestamp__date__gte=date_from, timestamp__date__lte=date_to)
        
        if movement_type:
            qs = qs.filter(movement_type=movement_type)
        
        if user_id:
            qs = qs.filter(user_id=user_id)
        
        return qs.order_by('-timestamp')

    def perform_create(self, serializer):
        with transaction.atomic():
            serializer.save(user=self.request.user)

    @action(detail=False, methods=['post'], url_path='add')
    def add_stock(self, request):
        product_id = request.data.get('product_id')
        quantity = request.data.get('quantity')
        notes = request.data.get('notes', '')
        
        if not product_id or not quantity:
            return Response({'error': 'product_id and quantity are required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            quantity = int(quantity)
            if quantity <= 0:
                return Response({'error': 'Quantity must be positive'}, status=status.HTTP_400_BAD_REQUEST)
        except ValueError:
            return Response({'error': 'Invalid quantity'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found'}, status=status.HTTP_404_NOT_FOUND)
        
        with transaction.atomic():
            # Get or create a default lot for this product
            lot = product.lots.filter(is_active=True).first()
            if not lot:
                # Create a default lot
                lot = ProductLot.objects.create(
                    product=product,
                    lot_number=f"LOT-{product.sku}-{timezone.now().strftime('%Y%m%d')}",
                    expiration_date=timezone.now().date() + timedelta(days=365),
                    current_quantity=0
                )
            
            lot.current_quantity += quantity
            lot.save()
            
            user = request.user
            movement = StockMovement.objects.create(
                product=product,
                lot=lot,
                user=user,
                user_full_name=f"{user.first_name} {user.last_name}".strip() or user.email,
                user_email=user.email,
                user_role=user.role if hasattr(user, 'role') else 'unknown',
                movement_type='STOCK_IN',
                quantity=quantity,
                notes=notes
            )
            
            # Create audit log
            create_audit_log(user, 'STOCK_IN', product, quantity, notes)
            
            serializer = self.get_serializer(movement)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='remove')
    def remove_stock(self, request):
        product_id = request.data.get('product_id')
        quantity = request.data.get('quantity')
        notes = request.data.get('notes', '')
        
        if not product_id or not quantity:
            return Response({'error': 'product_id and quantity are required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            quantity = int(quantity)
            if quantity <= 0:
                return Response({'error': 'Quantity must be positive'}, status=status.HTTP_400_BAD_REQUEST)
        except ValueError:
            return Response({'error': 'Invalid quantity'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if product.is_quarantined:
            return Response({'error': f'Product {product.sku} is currently in QUARANTINE and cannot be moved.'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            # Get active lots with stock
            lots = product.lots.filter(is_active=True, current_quantity__gt=0).order_by('expiration_date')
            
            if not lots.exists():
                return Response({'error': 'No stock available for this product'}, status=status.HTTP_400_BAD_REQUEST)
            
            total_available = sum(lot.current_quantity for lot in lots)
            if total_available < quantity:
                return Response({'error': f'Insufficient stock. Available: {total_available}'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Use FIFO - remove from oldest lots first
            remaining = quantity
            lot_used = None
            for lot in lots:
                if remaining <= 0:
                    break
                
                if lot.current_quantity >= remaining:
                    lot.current_quantity -= remaining
                    lot_used = lot
                    remaining = 0
                else:
                    remaining -= lot.current_quantity
                    lot.current_quantity = 0
                    lot_used = lot
                
                lot.save()
            
            user = request.user
            movement = StockMovement.objects.create(
                product=product,
                lot=lot_used,
                user=user,
                user_full_name=f"{user.first_name} {user.last_name}".strip() or user.email,
                user_email=user.email,
                user_role=user.role if hasattr(user, 'role') else 'unknown',
                movement_type='STOCK_OUT',
                quantity=quantity,
                notes=notes
            )
            
            # Create audit log
            create_audit_log(user, 'STOCK_OUT', product, quantity, notes)
            
            serializer = self.get_serializer(movement)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='quarantine')
    def quarantine_product(self, request):
        product_id = request.data.get('product_id')
        reason = request.data.get('reason', '')
        
        if not product_id:
            return Response({'error': 'product_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found'}, status=status.HTTP_404_NOT_FOUND)
        
        with transaction.atomic():
            product.is_quarantined = True
            product.quarantine_reason = reason
            product.quarantined_at = timezone.now()
            product.quarantined_by = request.user
            product.save()
            
            # Calculate total quantity
            total_qty = product.lots.filter(is_active=True).aggregate(Sum('current_quantity'))['current_quantity__sum'] or 0
            
            user = request.user
            movement = StockMovement.objects.create(
                product=product,
                user=user,
                user_full_name=f"{user.first_name} {user.last_name}".strip() or user.email,
                user_email=user.email,
                user_role=user.role if hasattr(user, 'role') else 'unknown',
                movement_type='QUARANTINED',
                quantity=total_qty,
                notes=reason
            )
            
            # Create audit log
            create_audit_log(user, 'QUARANTINED', product, total_qty, reason)
            
            serializer = self.get_serializer(movement)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='export/xlsx')
    def export_xlsx(self, request):
        if not OPENPYXL_AVAILABLE:
            return Response({'error': 'openpyxl non installata'}, status=500)
        
        movements = self.get_queryset()
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="movimenti_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Registro Movimenti'
        
        headers = ['Data e Ora', 'Prodotto', 'SKU', 'Tipo Azione', 'Quantità', 'Operatore', 'Ruolo', 'Note']
        ws.append(headers)
        
        header_fill = openpyxl.styles.PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for mov in movements:
            product_name = mov.product.name if mov.product else (mov.lot.product.name if mov.lot else 'N/A')
            product_sku = mov.product.sku if mov.product else (mov.lot.product.sku if mov.lot else 'N/A')
            
            ws.append([
                mov.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
                product_name,
                product_sku,
                mov.get_movement_type_display(),
                mov.quantity,
                mov.user_full_name or (mov.user.email if mov.user else 'N/A'),
                mov.user_role,
                mov.notes
            ])
        
        # Adjust Column Widths
        widths = [20, 45, 15, 15, 12, 25, 15, 50]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='export/pdf')
    def export_pdf(self, request):
        if not REPORTLAB_AVAILABLE:
            return Response({'error': 'reportlab non installata'}, status=500)
        
        movements = self.get_queryset()[:200]  # Increase limit slightly for better report
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="movimenti_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter, leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        header_style = styles['Heading1']
        header_style.textColor = colors.HexColor('#1e293b') # Slate-800
        
        title = Paragraph("Registro Movimenti di Magazzino", header_style)
        elements.append(title)
        elements.append(Paragraph(f"Generato il: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Cell style for wrapping
        cell_style = styles['Normal']
        cell_style.fontSize = 8
        cell_style.leading = 10
        
        data = [['Data/Ora', 'Prodotto', 'Tipo', 'Qtà', 'Operatore']]
        
        for mov in movements:
            product_name = mov.product.name if mov.product else (mov.lot.product.name if mov.lot else 'N/A')
            data.append([
                mov.timestamp.strftime('%d/%m/%Y %H:%M'),
                Paragraph(product_name, cell_style),
                mov.get_movement_type_display(),
                str(mov.quantity),
                Paragraph(mov.user_full_name or mov.user_email or 'N/A', cell_style)
            ])
        
        table = Table(data, colWidths=[1.2*inch, 3.2*inch, 1.2*inch, 0.7*inch, 1.2*inch])
        
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]
        
        # Zebra coloring
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f1f5f9')))
        
        table.setStyle(TableStyle(table_style))
        
        elements.append(table)
        doc.build(elements)
        return response
