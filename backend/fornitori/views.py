from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.db.models import Sum
from core.permissions import IsMagazziniere, IsAmministratore
from .models import Supplier

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from .serializers import SupplierSerializer

class SupplierViewSet(viewsets.ModelViewSet):
    serializer_class = SupplierSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'export_catalog_pdf', 'export_catalog_xlsx']:
            return [IsMagazziniere()]
        return [IsAmministratore()]

    def get_queryset(self):
        return Supplier.objects.filter(is_active=True)

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()

    @action(detail=True, methods=['get'], url_path='export-catalog/xlsx')
    def export_catalog_xlsx(self, request, pk=None):
        if not OPENPYXL_AVAILABLE:
            return Response({'error': 'openpyxl non installata'}, status=500)
        
        supplier = self.get_object()
        products = supplier.products.filter(is_active=True).distinct()
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="catalogo_{supplier.name.replace(" ", "_")}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Catalogo Fornitore'
        
        headers = ['SKU', 'Nome Prodotto', 'Categoria', 'Prezzo Unitario', 'Stock Attuale']
        ws.append(headers)
        
        # Header Styling
        header_fill = openpyxl.styles.PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
        header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        for product in products:
            total_qty = product.lots.filter(is_active=True).aggregate(Sum('current_quantity'))['current_quantity__sum'] or 0
            ws.append([
                product.sku,
                product.name,
                product.category.name if product.category else 'N/D',
                float(product.unit_price),
                f"{total_qty} {product.unit_of_measure}"
            ])
        
        # Column Widths
        widths = [15, 45, 25, 15, 15]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
        wb.save(response)
        return response

    @action(detail=True, methods=['get'], url_path='export-catalog/pdf')
    def export_catalog_pdf(self, request, pk=None):
        if not REPORTLAB_AVAILABLE:
            return Response({'error': 'reportlab non installata'}, status=500)
            
        supplier = self.get_object()
        products = supplier.products.filter(is_active=True).distinct()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="catalogo_{supplier.name.replace(" ", "_")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter, leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        header_style = styles['Heading1']
        header_style.textColor = colors.HexColor('#6366f1') # Indigo-500
        
        title = Paragraph(f"Catalogo Fornitore: {supplier.name}", header_style)
        elements.append(title)
        elements.append(Paragraph(f"Data Generazione: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Cell style for wrapping
        cell_style = styles['Normal']
        cell_style.fontSize = 8
        
        data = [['SKU', 'Prodotto', 'Categoria', 'Prezzo', 'Stock Magazzino']]
        
        for product in products:
            total_qty = product.lots.filter(is_active=True).aggregate(Sum('current_quantity'))['current_quantity__sum'] or 0
            data.append([
                product.sku,
                Paragraph(product.name, cell_style),
                product.category.name if product.category else 'N/D',
                f"€{product.unit_price}",
                f"{total_qty} {product.unit_of_measure}"
            ])
            
        table = Table(data, colWidths=[1.2*inch, 2.8*inch, 1.5*inch, 1*inch, 1.2*inch])
        
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]
        
        # Zebra
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f5f3ff')))
        
        table.setStyle(TableStyle(table_style))
        
        elements.append(table)
        doc.build(elements)
        return response