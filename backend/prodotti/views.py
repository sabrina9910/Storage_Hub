from django.http import HttpResponse

from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum, Q
from django.db import transaction
from django.utils import timezone
from datetime import timedelta
import csv

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import xml.etree.ElementTree as ET
    from xml.dom import minidom
    XML_AVAILABLE = True
except ImportError:
    XML_AVAILABLE = False

from core.permissions import IsMagazziniere, IsAmministratore
from .models import Product, ProductLot
from .serializers import ProductSerializer, ProductLotSerializer
from .filters import ProductFilter
from auditlog.utils import create_audit_log
from movimenti.models import StockMovement

class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    filterset_class = ProductFilter

    def get_permissions(self):
        if self.action in [
            'list', 'retrieve', 'alerts', 'restore_quarantine', 
            'export_xlsx', 'export_pdf', 'export_xml', 'export_csv',
            'export_catalog_xlsx', 'export_catalog_pdf', 'import_catalog_xlsx'
        ]:
            return [IsMagazziniere()]
        return [IsAmministratore()]


    def get_queryset(self):
        user = self.request.user
        if not user or not user.is_authenticated:
            return Product.objects.none()
        
        # Show all products for superuser
        if getattr(user, 'is_superuser', False):
            return Product.objects.all().select_related('category').prefetch_related('suppliers').distinct()
        return Product.objects.filter(is_active=True).select_related('category').prefetch_related('suppliers').distinct()

    def perform_create(self, serializer):
        serializer.save(owner=self.request.user)

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def alerts(self, request):
        today = timezone.now().date()
        seven_days_later = today + timedelta(days=7)

        products = Product.objects.filter(is_active=True).annotate(
            total_stock=Sum('lots__current_quantity', filter=Q(lots__is_active=True))
        )
        low_stock_products = []
        for p in products:
            qty = p.total_stock or 0
            if qty < p.min_stock_threshold:
                low_stock_products.append({
                    "id": p.id,
                    "name": p.name,
                    "sku": p.sku,
                    "current_stock": qty,
                    "threshold": p.min_stock_threshold
                })

        expiring_lots_qs = ProductLot.objects.filter(
            is_active=True,
            current_quantity__gt=0,
            expiration_date__lte=seven_days_later,
            expiration_date__gte=today
        ).select_related('product')
        
        expired_lots_qs = ProductLot.objects.filter(
            is_active=True,
            current_quantity__gt=0,
            expiration_date__lt=today
        ).select_related('product')

        expiring_lots = [
            {
                "id": lot.id,
                "product_name": lot.product.name,
                "lot_number": lot.lot_number,
                "quantity": lot.current_quantity,
                "expiration_date": lot.expiration_date
            } for lot in expiring_lots_qs
        ]
        
        expired_lots = [
            {
                "id": lot.id,
                "product_name": lot.product.name,
                "lot_number": lot.lot_number,
                "quantity": lot.current_quantity,
                "expiration_date": lot.expiration_date
            } for lot in expired_lots_qs
        ]

        return Response({
            "low_stock": low_stock_products,
            "expiring_soon": expiring_lots,
            "already_expired": expired_lots,
            "total_inventory_value": sum(
                [(p.total_stock or 0) * p.unit_price for p in products]
            )
        })

    @action(detail=True, methods=['get'], url_path='export/csv')
    def export_csv(self, request, pk=None):
        product = self.get_object()
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="product_{product.sku}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Product Name', 'SKU', 'Quantity', 'Supplier', 'Price', 'Last Update'])
        
        total_qty = product.lots.filter(is_active=True).aggregate(Sum('current_quantity'))['current_quantity__sum'] or 0
        supplier_names = ', '.join([s.name for s in product.suppliers.all()])
        
        writer.writerow([
            product.name,
            product.sku,
            total_qty,
            supplier_names or 'N/A',
            f"{product.unit_price}",
            timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        ])
        return response

    @action(detail=True, methods=['get'], url_path='export/xlsx')
    def export_xlsx(self, request, pk=None):
        if not OPENPYXL_AVAILABLE:
            return Response({'error': 'openpyxl non installata'}, status=500)
        
        product = self.get_object()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="product_{product.sku}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Dettagli Prodotto'
        
        headers = ['Nome Prodotto', 'SKU', 'Giacenza', 'Fornitori', 'Prezzo', 'Ultimo Aggiornamento']
        ws.append(headers)
        
        # Header Styling
        fill = openpyxl.styles.PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
        font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        total_qty = product.lots.filter(is_active=True).aggregate(Sum('current_quantity'))['current_quantity__sum'] or 0
        supplier_names = ', '.join([s.name for s in product.suppliers.all()])
        
        ws.append([
            product.name,
            product.sku,
            f"{total_qty} {product.unit_of_measure}",
            supplier_names or 'N/A',
            float(product.unit_price),
            timezone.now().strftime('%d/%m/%Y %H:%M')
        ])
        
        # Adjust Column Widths
        column_widths = [40, 15, 15, 40, 12, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
        wb.save(response)
        return response

    @action(detail=True, methods=['get'], url_path='export/xml')
    def export_xml(self, request, pk=None):
        if not XML_AVAILABLE:
            return Response({'error': 'xml libraries not available'}, status=500)
        
        product = self.get_object()
        response = HttpResponse(content_type='application/xml')
        response['Content-Disposition'] = f'attachment; filename="product_{product.sku}.xml"'
        
        root = ET.Element('product')
        ET.SubElement(root, 'name').text = product.name
        ET.SubElement(root, 'sku').text = product.sku
        
        total_qty = product.lots.filter(is_active=True).aggregate(Sum('current_quantity'))['current_quantity__sum'] or 0
        ET.SubElement(root, 'quantity').text = str(total_qty)
        
        supplier_names = ', '.join([s.name for s in product.suppliers.all()])
        ET.SubElement(root, 'supplier').text = supplier_names or 'N/A'
        ET.SubElement(root, 'price').text = str(product.unit_price)
        ET.SubElement(root, 'last_update').text = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        
        xml_str = minidom.parseString(ET.tostring(root)).toprettyxml(indent='  ')
        response.write(xml_str)
        return response

    @action(detail=True, methods=['get'], url_path='export/pdf')
    def export_pdf(self, request, pk=None):
        if not REPORTLAB_AVAILABLE:
            return Response({'error': 'reportlab non installata'}, status=500)
        
        product = self.get_object()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="product_{product.sku}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter, leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Header Style
        header_style = styles['Heading1']
        header_style.textColor = colors.HexColor('#4f46e5')  # Indigo-600
        
        title = Paragraph(f"Dettaglio Prodotto: {product.name}", header_style)
        elements.append(title)
        elements.append(Paragraph(f"Data Esportazione: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        total_qty = product.lots.filter(is_active=True).aggregate(Sum('current_quantity'))['current_quantity__sum'] or 0
        supplier_names = ', '.join([s.name for s in product.suppliers.all()])
        
        # Use Paragraphs for wrapping in table cells
        cell_style = styles['Normal']
        cell_style.fontSize = 9
        
        data = [
            ['SKU', 'Prodotto', 'Giacenza', 'Fornitori', 'Prezzo'],
            [
                product.sku,
                Paragraph(product.name, cell_style),
                f"{total_qty} {product.unit_of_measure}",
                Paragraph(supplier_names or 'N/A', cell_style),
                f"€ {product.unit_price}"
            ]
        ]
        
        table = Table(data, colWidths=[1.0*inch, 3.0*inch, 1.2*inch, 1.3*inch, 1.0*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        elements.append(table)
        doc.build(elements)
        return response

    @action(detail=False, methods=['get'], url_path='export-catalog/xlsx')
    def export_catalog_xlsx(self, request):
        if not OPENPYXL_AVAILABLE:
            return Response({'error': 'openpyxl non installata'}, status=500)
        
        products = self.filter_queryset(self.get_queryset())
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="catalogo_prodotti.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Catalogo Prodotti'
        
        headers = ['SKU', 'Nome', 'Categoria', 'Prezzo Unitario', 'Quantità Totale', 'Unità di Misura']
        ws.append(headers)
        
        # Header Styling
        header_fill = openpyxl.styles.PatternFill(start_color="0ea5e9", end_color="0ea5e9", fill_type="solid")
        header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for product in products:
            total_qty = product.lots.filter(is_active=True).aggregate(Sum('current_quantity'))['current_quantity__sum'] or 0
            ws.append([
                product.sku,
                product.name,
                product.category.name if product.category else 'N/D',
                float(product.unit_price),
                total_qty,
                product.unit_of_measure
            ])
        
        # Adjust column widths automatically
        widths = [15, 45, 25, 15, 15, 15]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='export-catalog/pdf')
    def export_catalog_pdf(self, request):
        if not REPORTLAB_AVAILABLE:
            return Response({'error': 'reportlab non installata'}, status=500)
            
        products = self.filter_queryset(self.get_queryset())
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="catalogo_prodotti.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter, leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        header_style = styles['Heading1']
        header_style.textColor = colors.HexColor('#0ea5e9') # Sky-500
        
        title = Paragraph("Catalogo Prodotti - StorageHub", header_style)
        elements.append(title)
        elements.append(Paragraph(f"Data Generazione: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Cell style for wrapping
        cell_style = styles['Normal']
        cell_style.fontSize = 8
        cell_style.leading = 10
        
        data = [['SKU', 'Prodotto', 'Categoria', 'Prezzo', 'Stock']]
        
        for i, product in enumerate(products):
            total_qty = product.lots.filter(is_active=True).aggregate(Sum('current_quantity'))['current_quantity__sum'] or 0
            data.append([
                product.sku,
                Paragraph(product.name, cell_style),
                product.category.name if product.category else 'N/D',
                f"€{product.unit_price}",
                f"{total_qty} {product.unit_of_measure}"
            ])
            
        table = Table(data, colWidths=[1.2*inch, 2.5*inch, 1.5*inch, 1*inch, 1.3*inch])
        
        # Table Styling
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0ea5e9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]
        
        # Alternate row coloring (Zebra)
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8fafc')))
        
        table.setStyle(TableStyle(table_style))
        
        elements.append(table)
        doc.build(elements)
        return response

    @action(detail=False, methods=['post'], url_path='import-catalog/xlsx')
    def import_catalog_xlsx(self, request):
        if not OPENPYXL_AVAILABLE:
            return Response({'error': 'openpyxl non installata'}, status=500)
            
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Nessun file caricato'}, status=400)
            
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) < 2:
                return Response({'error': 'Il file è troppo corto o vuoto'}, status=400)
                
            headers = [str(h).lower().strip() for h in rows[0]]
            # nome, sku, categoria, prezzo, um, soglia
            
            import_count = 0
            errors = []
            
            with transaction.atomic():
                from .models import ProductCategory
                for i, row in enumerate(rows[1:], start=2):
                    row_data = dict(zip(headers, row))
                    
                    sku = str(row_data.get('sku', '') or '').strip()
                    name = str(row_data.get('nome', '') or '').strip()
                    cat_name = str(row_data.get('categoria', '') or '').strip()
                    price_val = row_data.get('prezzo', 0)
                    um = str(row_data.get('um', 'pz') or 'pz').strip()
                    threshold_val = row_data.get('soglia', 10)
                    
                    if not sku or not name:
                        errors.append(f"Riga {i}: SKU e Nome sono obbligatori")
                        continue
                        
                    category = None
                    if cat_name:
                        category, _ = ProductCategory.objects.get_or_create(name=cat_name)
                    
                    try:
                        price = float(price_val) if price_val is not None else 0.0
                    except:
                        price = 0.0
                        
                    try:
                        threshold = int(threshold_val) if threshold_val is not None else 10
                    except:
                        threshold = 10
                        
                    product, created = Product.objects.update_or_create(
                        sku=sku,
                        defaults={
                            'name': name,
                            'category': category,
                            'unit_price': price,
                            'unit_of_measure': um,
                            'min_stock_threshold': threshold,
                            'owner': request.user,
                            'is_active': True
                        }
                    )
                    import_count += 1
                    
            return Response({
                'message': f'Importati/Aggiornati {import_count} prodotti',
                'errors': errors
            })
            
        except Exception as e:
            return Response({'error': str(e)}, status=500)


    @action(detail=True, methods=['patch'], url_path='blacklist')
    def blacklist(self, request, pk=None):
        self.permission_classes = [IsAmministratore]
        self.check_permissions(request)
        
        product = self.get_object()
        reason = request.data.get('reason', '')
        
        product.is_blacklisted = True
        product.blacklist_reason = reason
        product.save()
        
        # Create audit log
        create_audit_log(request.user, 'BLACKLISTED', product, 0, reason)
        
        serializer = self.get_serializer(product)
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='restore')
    def restore(self, request, pk=None):
        self.permission_classes = [IsAmministratore]
        self.check_permissions(request)
        
        product = self.get_object()
        
        product.is_blacklisted = False
        product.blacklist_reason = ''
        product.save()
        
        # Create audit log
        create_audit_log(request.user, 'RESTORED', product, 0, 'Product restored from blacklist')
        
        serializer = self.get_serializer(product)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='blacklisted')
    def blacklisted(self, request):
        self.permission_classes = [IsAmministratore]
        self.check_permissions(request)
        
        blacklisted_products = Product.objects.filter(is_blacklisted=True)
        serializer = self.get_serializer(blacklisted_products, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['patch'], url_path='restore-quarantine')
    def restore_quarantine(self, request, pk=None):
        product = self.get_object()
        
        with transaction.atomic():
            product.is_quarantined = False
            product.quarantine_reason = ''
            product.quarantined_at = None
            product.quarantined_by = None
            product.save()
            
            # Calculate total quantity
            total_qty = product.lots.filter(is_active=True).aggregate(Sum('current_quantity'))['current_quantity__sum'] or 0
            
            user = request.user
            StockMovement.objects.create(
                product=product,
                user=user,
                user_full_name=f"{user.first_name} {user.last_name}".strip() or user.email,
                user_email=user.email,
                user_role=user.role if hasattr(user, 'role') else 'unknown',
                movement_type='RESTORED',
                quantity=total_qty,
                notes="Ripristinato dalla quarantena"
            )
            
            # Create audit log
            create_audit_log(user, 'RESTORED', product, total_qty, "Restore from quarantine")
            
        serializer = self.get_serializer(product)
        return Response(serializer.data)

class ProductLotViewSet(viewsets.ModelViewSet):
    serializer_class = ProductLotSerializer
    permission_classes = [IsMagazziniere]

    def get_queryset(self):
        user = self.request.user
        if not user or not user.is_authenticated:
            return ProductLot.objects.none()
            
        if getattr(user, 'is_superuser', False):
            qs = ProductLot.objects.all()
        else:
            qs = ProductLot.objects.filter(is_active=True)
            
        return qs.select_related('product', 'product__category').prefetch_related('product__suppliers').order_by('expiration_date')

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()
